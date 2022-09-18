import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pyodbc as py
from wejscie import df
import pandas as pd
import pywin as py
from win32 import win32file


name = 'sample_book.xlsx'
path = 'C:\\Users\\user\\Documents\\excellookform\\'
wb = Workbook()


ws1 = wb.active
# ws1.title = 'Produkcja'

# wb.save(filename = path + name)
licznik = 0

for x in df:
    licznik += 1
    # print(x)


data_frame = pd.DataFrame(data = df)
# print(data_frame)
data_frame.to_excel(path + name)
work = load_workbook(path + name)
ws = work.active
ws.auto_filter.ref = "A1:Z" + str(licznik)


ws.title = 'Produkcja'


#
work.save(path+name)
#
